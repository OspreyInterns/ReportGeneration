
import sqlite3 as sqlite
import openpyxl
from openpyxl.styles import Alignment, PatternFill

# Reads from the injection table to sum up the injections


def straight_to_patient(case_number, file_name):

    _con = sqlite.connect(file_name)
    mismatch = False

    with _con:

        contrast_inj = 0.
        alt_contrast_inj = 0
        _cur = _con.cursor()
        _cur.execute('SELECT * FROM CMSWInjections')

        _rows = _cur.fetchall()

        for _row in _rows:
            if _row[1] == case_number and _row[5] == 1 and _row[18] == 0:
                contrast_inj += _row[20]
                if _row[17] != 0 and mismatch is False:
                    print('Case', _row[1], 'contains a mismatch between % and volume diverted')
                    mismatch = True
            if _row[1] == case_number and _row[5] == 1 and _row[17] == 0:
                alt_contrast_inj += _row[20]
                if round(_row[12], 4) != round(_row[16] + _row[19], 4) and _row[30] == 0 and _row[32] == 0:
                    if _row[29] != 0:
                        print('Injection', _row[0], 'suspicious', _row[12], '!=', _row[16] + _row[19])

        return [contrast_inj, alt_contrast_inj]


def excel_write(file_name, cmsw):

    con = sqlite.connect(file_name)

    with con:

        cur = con.cursor()
        cur.execute('SELECT * FROM CMSWCases')
        rows = cur.fetchall()
        check_cases = [('Case ID/Patient ID Field #', )]

        for row in rows:
            to_patient = straight_to_patient(row[0], file_name)
            if row[8] == 0:
                perc_threshold = 'N/A'
            else:
                perc_threshold = row[15]/row[8]*100
            if row[2] == '2.1.24':
                check_cases.append(('', '', '', row[5][0:10], row[1][-12:-1], '', row[19], row[8], row[13],
                                   row[14], row[15], row[16], perc_threshold, '', to_patient[0], '', '', to_patient[1],
                                   to_patient[0] - to_patient[1]))
            else:
                check_cases.append(('', '', '', row[5][0:10], row[1][-12:-1], row[20][-8:-1], row[19], row[8], row[13],
                                   row[14], row[15], row[16], perc_threshold, '', to_patient[0], '', '', to_patient[1],
                                   to_patient[0]-to_patient[1]))

    xlsx_name = cmsw + 'DyeMinishOutput.xlsx'
    wb = openpyxl.load_workbook('F173-A_template-DyeMINISH Display Data Summary.xlsx')
    data_sheet = wb.active
    data_sheet.title = 'Sheet1'

    for row in range(len(check_cases)):
        for col in range(len(check_cases[row])):
            if row != 0 and float(check_cases[row][6]) <= 5.:
                data_sheet.cell(row=row + 1, column=col + 1, value=check_cases[row][col]).fill = PatternFill(
                    fill_type="solid", start_color='FFFF00', end_color='FFFF00')
                data_sheet.cell(row=row + 1, column=16, value='Case less than 5 Minutes')
            elif row != 0 and check_cases[row][8] == 0 and check_cases[row][9] == 0 and check_cases[row][10] == 0 \
                    and check_cases[row][11] == 0:
                data_sheet.cell(row=row + 1, column=col + 1, value=check_cases[row][col]).fill = PatternFill(
                    fill_type="solid", start_color='FFFF00', end_color='FFFF00')
                data_sheet.cell(row=row + 1, column=16, value='No contrast injected')
            else:
                data_sheet.cell(row=row + 1, column=col + 1, value=check_cases[row][col])
            data_sheet.cell(row=row + 1, column=col + 1).alignment = Alignment(wrapText=True)

    wb.save(xlsx_name)

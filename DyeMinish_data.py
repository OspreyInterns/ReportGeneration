import sqlite3 as sqlite
import logging
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
import datetime
import math
from openpyxl.chart import (
    LineChart,
    Reference)
# case column numbers

CMSW_CASE_ID = 0
CASE_ID = 1
SERIAL_NUMBER = 3
DATE_OF_PROCEDURE = 5
DYEVERT_USED = 6
THRESHOLD_VOLUME = 8
ATTEMPTED_CONTRAST_INJECTION_VOLUME = 13
DIVERTED_CONTRAST_VOLUME = 14
CUMULATIVE_VOLUME_TO_PATIENT = 15
PERCENTAGE_CONTRAST_DIVERTED = 16
DYEVERTEZ = 23
FLOW_RATE_TO_FROM_SYRINGE = 28
FLOW_RATE_TO_PATIENT = 29
#
# # Use these for iPad
# TOTAL_DURATION = 20
# END_TIME = 19
# # (from injection table)
# IS_AN_INJECTION = 8
# LINEAR_DYEVERT_MOVEMENT = 18
# DYEVERT_CONTRAST_VOLUME_DIVERTED = 20
# PERCENT_CONTRAST_SAVED = 21
# CONTRAST_VOLUME_TO_PATIENT = 23
# PREDOMINANT_CONTRAST_LINE_PRESSURE = 34


# Use these for CMSW
TOTAL_DURATION = 19
END_TIME = 20
# (from injection table)
IS_AN_INJECTION = 5
LINEAR_DYEVERT_MOVEMENT = 15
DYEVERT_CONTRAST_VOLUME_DIVERTED = 17
PERCENT_CONTRAST_SAVED = 18
CONTRAST_VOLUME_TO_PATIENT = 20
PREDOMINANT_CONTRAST_LINE_PRESSURE = 30

# Highlight color
YELLOW = 'FFFF00'


# Reads from the injection table to sum up the injections


def straight_to_patient(file_name):
    con = sqlite.connect(file_name)
    #with con:

        #missing_summary = []
        #cur = con.cursor()
        #cur.execute('SELECT * FROM CMSWCases')
        #rows = cur.fetchall()

        #for row in rows:
            #if row[ATTEMPTED_CONTRAST_INJECTION_VOLUME] == row[DIVERTED_CONTRAST_VOLUME] == \
                    #row[CUMULATIVE_VOLUME_TO_PATIENT] == row[DIVERTED_CONTRAST_VOLUME] == 0:
                #missing_summary.append(row[CASE_ID])
                #logging.warning('Case ' + str(row[CASE_ID]) + ' may be missing summary data')

    with con:

        contrast_inj = []
        cur = con.cursor()
        cur.execute('SELECT * FROM CMSWInjections')

        rows = cur.fetchall()

        for placeholder in range(len(rows)):
            contrast_inj.append([0, 0])
        for row in rows:
            #if row[CMSW_CASE_ID] not in missing_summary:
            if row[IS_AN_INJECTION] == 1 and (row[PERCENT_CONTRAST_SAVED] <= 1 or row[LINEAR_DYEVERT_MOVEMENT] <= 20) \
                    and row[PREDOMINANT_CONTRAST_LINE_PRESSURE] == 0 and row[FLOW_RATE_TO_PATIENT] > .5:
                contrast_inj[row[CASE_ID] - 1][0] += row[CONTRAST_VOLUME_TO_PATIENT]
            if row[IS_AN_INJECTION] == 1 and (row[DYEVERT_CONTRAST_VOLUME_DIVERTED] == 0
                                              or row[LINEAR_DYEVERT_MOVEMENT] <= 20):
                contrast_inj[row[CASE_ID] - 1][1] += row[CONTRAST_VOLUME_TO_PATIENT]

        return contrast_inj


def would_be_saved(file_name):
    con = sqlite.connect(file_name)

    with con:

        cur = con.cursor()
        cur.execute('SELECT * FROM CMSWCases')
        rows = cur.fetchall()
        case_info = []

        for row in rows:
            case_info.append([row[CMSW_CASE_ID], row[ATTEMPTED_CONTRAST_INJECTION_VOLUME],
                              row[THRESHOLD_VOLUME], row[CUMULATIVE_VOLUME_TO_PATIENT], row[SERIAL_NUMBER]])

            print(row, case_info)
    what_if = []

    direct_injected = straight_to_patient(file_name)

    print("now printing case info, case by case")
    for case in case_info:
        print(case)
        vol_inj_off = direct_injected[case[0] - 1][0]
        #        print(" case 3",case[3]," vol-inj-off",vol_inj_off)
        vol_inj_on = case[3] - vol_inj_off
        vol_att_on = case[1] - vol_inj_off
        if vol_att_on != 0:
            perc_savings_on = 1. - (vol_inj_on / vol_att_on)
            would_be_total = vol_inj_on + (vol_inj_off * (1. - perc_savings_on))
            if case[2] != 0:
                would_be_portion = (would_be_total / case[2]) * 100
            else:
                debug_msg = 'CMSW, ' + str(case[4]) + ' case ' + str(case[0]) + ' has zero threshold'
                logging.warning(debug_msg)
                would_be_portion = 0
            what_if.append([would_be_total, would_be_portion])
        elif case[2] == 0 and vol_att_on == 0:
            debug_msg = 'CMSW, ' + str(case[4]) + ' case ' + str(case[0]) + ' has zero threshold'
            logging.warning(debug_msg)
            what_if.append([vol_inj_off, 0])
        else:
            what_if.append([vol_inj_off, vol_inj_off / case[2]])

    return what_if


def list_builder(file_names):
    """Takes the list of files and builds the list of lists to write"""
    DyeMinishCases = []

    for file_name in file_names:
        con = sqlite.connect(file_name)
        with con:
            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWCases')
            rows = cur.fetchall()
        global TOTAL_DURATION, END_TIME, IS_AN_INJECTION, LINEAR_DYEVERT_MOVEMENT, DYEVERT_CONTRAST_VOLUME_DIVERTED, \
            DYEVERT_CONTRAST_VOLUME_DIVERTED, PERCENT_CONTRAST_SAVED, CONTRAST_VOLUME_TO_PATIENT, \
            PREDOMINANT_CONTRAST_LINE_PRESSURE, FLOW_RATE_TO_FROM_SYRINGE, FLOW_RATE_TO_PATIENT
        if not rows == []:
            if rows[0][2] == '2.1.56' or rows[0][2] == '2.1.24' or rows[0][2] == '2.1.67':
                TOTAL_DURATION = 19
                END_TIME = 20
                IS_AN_INJECTION = 5
                LINEAR_DYEVERT_MOVEMENT = 15
                DYEVERT_CONTRAST_VOLUME_DIVERTED = 17
                PERCENT_CONTRAST_SAVED = 18
                CONTRAST_VOLUME_TO_PATIENT = 20
                FLOW_RATE_TO_FROM_SYRINGE = 28
                FLOW_RATE_TO_PATIENT = 29
                PREDOMINANT_CONTRAST_LINE_PRESSURE = 30
            else:
                TOTAL_DURATION = 20
                END_TIME = 19
                IS_AN_INJECTION = 8
                LINEAR_DYEVERT_MOVEMENT = 18
                DYEVERT_CONTRAST_VOLUME_DIVERTED = 20
                PERCENT_CONTRAST_SAVED = 21
                CONTRAST_VOLUME_TO_PATIENT = 23
                FLOW_RATE_TO_FROM_SYRINGE = 32
                FLOW_RATE_TO_PATIENT = 33
                PREDOMINANT_CONTRAST_LINE_PRESSURE = 34
            what_if = would_be_saved(file_name)
            to_patient = straight_to_patient(file_name)
            for row in rows:
                if row[DYEVERT_USED] == 1:
                    if row[DYEVERTEZ] == 0:
                        pmdv = 'DV'
                    else:
                        pmdv = 'DVEZ'
                else:
                    pmdv = 'PM'
                if row[THRESHOLD_VOLUME] == 0:
                    perc_threshold = 'N/A'
                else:
                    perc_threshold = row[CUMULATIVE_VOLUME_TO_PATIENT] / row[THRESHOLD_VOLUME] * 100
                print(row[2], row[END_TIME])
                if row[2] == '2.1.21' or row[2] == '2.1.24' or row[2] == '2.0.1981' or row[2] == '2.0.2013':
                    DyeMinishCases.append((row[DATE_OF_PROCEDURE][0:10], row[CASE_ID][-12:-4], row[TOTAL_DURATION],
                                           row[THRESHOLD_VOLUME], row[ATTEMPTED_CONTRAST_INJECTION_VOLUME],
                                           row[DIVERTED_CONTRAST_VOLUME], row[CUMULATIVE_VOLUME_TO_PATIENT],
                                           row[PERCENTAGE_CONTRAST_DIVERTED], perc_threshold,
                                           to_patient[row[CMSW_CASE_ID] - 1][0], what_if[row[CMSW_CASE_ID] - 1][0],
                                           what_if[row[CMSW_CASE_ID] - 1][1], row[SERIAL_NUMBER], pmdv,
                                           row[CMSW_CASE_ID], row[END_TIME][-12:-4]))
                else:
                    # put end time into datetime object.
                    if row[2] == '2.2.44':
                        case_end = datetime.datetime.strptime(row[END_TIME], '%Y/%m/%d %I:%M.%S %p')
                    elif row[2] == '2.2.38':
                        case_end = datetime.datetime.strptime(row[END_TIME], '%m/%d/%y %I:%M %p')
                    elif row[2] == '2.1.56':
                        case_end = datetime.datetime.strptime(row[END_TIME], '%d %b %Y %H:%M:%S')
                    else:
                        # else this is 2.1.67 system.
                        case_end = datetime.datetime.strptime(row[END_TIME], '%Y-%m-%d %H:%M:%S.%f')
                    print(row[2], case_end.strftime('%H:%M:%S'))
                    case_end = case_end.strftime('%H:%M:%S')
                    DyeMinishCases.append((row[DATE_OF_PROCEDURE][0:10], row[CASE_ID][-12:-4],
                                           # case_end.strftime('%H:%M:%S'),
                                           row[TOTAL_DURATION], row[THRESHOLD_VOLUME],
                                           # row[END_TIME], row[TOTAL_DURATION], row[THRESHOLD_VOLUME],
                                           row[ATTEMPTED_CONTRAST_INJECTION_VOLUME], row[DIVERTED_CONTRAST_VOLUME],
                                           row[CUMULATIVE_VOLUME_TO_PATIENT], row[PERCENTAGE_CONTRAST_DIVERTED],
                                           perc_threshold, to_patient[row[CMSW_CASE_ID] - 1][0],
                                           what_if[row[CMSW_CASE_ID] - 1][0], what_if[row[CMSW_CASE_ID] - 1][1],
                                           row[SERIAL_NUMBER], pmdv, row[CMSW_CASE_ID], case_end))

    return DyeMinishCases


def dyevert_uses(file_names):
    """Connects to an individual database to calculate the volume of contrast injected
    and the number of times contrast was injected both in puffs and injections
    Volume data is currently unused
    Also adds tracking for first and last injection times
    """
    uses = []
    dyevert_used_inj = 0
    dyevert_not_used_inj = 0
    dyevert_used_puff = 0
    dyevert_not_used_puff = 0
    vol_used_inj = 0
    vol_not_used_inj = 0
    vol_used_puff = 0
    vol_not_used_puff = 0
    last_inj = ''
    first_inj = ''
    number_of_cases = 0
    for file_name in file_names:
        line = 0
        con = sqlite.connect(file_name)
        with con:
            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWCases')
            rows = cur.fetchall()
            number_of_cases += len(rows)
        global TOTAL_DURATION, END_TIME, IS_AN_INJECTION, LINEAR_DYEVERT_MOVEMENT, DYEVERT_CONTRAST_VOLUME_DIVERTED, \
            DYEVERT_CONTRAST_VOLUME_DIVERTED, PERCENT_CONTRAST_SAVED, CONTRAST_VOLUME_TO_PATIENT, \
            PREDOMINANT_CONTRAST_LINE_PRESSURE, FLOW_RATE_TO_FROM_SYRINGE, FLOW_RATE_TO_PATIENT
        if not rows == []:
            if rows[0][2] == '2.1.56' or rows[0][2] == '2.1.24' or rows[0][2] == '2.1.67':
                TOTAL_DURATION = 19
                END_TIME = 20
                IS_AN_INJECTION = 5
                LINEAR_DYEVERT_MOVEMENT = 15
                DYEVERT_CONTRAST_VOLUME_DIVERTED = 17
                PERCENT_CONTRAST_SAVED = 18
                CONTRAST_VOLUME_TO_PATIENT = 20
                FLOW_RATE_TO_FROM_SYRINGE = 28
                FLOW_RATE_TO_PATIENT = 29
                PREDOMINANT_CONTRAST_LINE_PRESSURE = 30
            else:
                TOTAL_DURATION = 20
                END_TIME = 19
                IS_AN_INJECTION = 8
                LINEAR_DYEVERT_MOVEMENT = 18
                DYEVERT_CONTRAST_VOLUME_DIVERTED = 20
                PERCENT_CONTRAST_SAVED = 21
                CONTRAST_VOLUME_TO_PATIENT = 23
                FLOW_RATE_TO_FROM_SYRINGE = 32
                FLOW_RATE_TO_PATIENT = 33
                PREDOMINANT_CONTRAST_LINE_PRESSURE = 34
        with con:

            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWInjections')
            rows = cur.fetchall()
            for row in rows:
                if row[CASE_ID] != line:
                    if first_inj == '':  #this sets the first first injection time
                        first_inj = rows[0][2][-12:-4]
                    uses.append([dyevert_not_used_inj, dyevert_used_inj, dyevert_not_used_puff,   #hre we write in the numbers we have counted from the loops once we detect that we're on a new case
                                dyevert_used_puff, first_inj, last_inj])
                    dyevert_used_inj = 0
                    dyevert_not_used_inj = 0
                    dyevert_used_puff = 0
                    dyevert_not_used_puff = 0
                    vol_used_inj = 0
                    vol_not_used_inj = 0
                    vol_used_puff = 0
                    vol_not_used_puff = 0
                    first_inj = row[2][-12:-4]
                    line += 1
                    while line < row[CASE_ID]:  #this handles cases which have no injections by adding appropriate blank data to maintain proper spacing
                        uses.append([0, 0, 0, 0, '', ''])
                        line += 1
                if row[CASE_ID] == line:
                    last_inj = row[2][-12:-4]
                    if row[CONTRAST_VOLUME_TO_PATIENT] + row[DYEVERT_CONTRAST_VOLUME_DIVERTED] >= 3: #1 is an injection, 2 is a puff
                        puff_inj = 1
                    elif row[CONTRAST_VOLUME_TO_PATIENT] + row[DYEVERT_CONTRAST_VOLUME_DIVERTED] <= 2:
                        puff_inj = 2
                    elif row[FLOW_RATE_TO_FROM_SYRINGE] >= 2.5:
                        puff_inj = 1
                    elif row[FLOW_RATE_TO_FROM_SYRINGE] <= 2:
                        puff_inj = 2
                    if round(row[FLOW_RATE_TO_FROM_SYRINGE], 2) != 0 and round(row[FLOW_RATE_TO_PATIENT], 2) != 0: #here it decides if the DyeVert was used, based on whether or not contrast was diverted
                        if row[IS_AN_INJECTION] == 1 and row[PERCENT_CONTRAST_SAVED] == 0 and puff_inj == 1: #it also sums up the contrast used in each category, though this is not currently used
                            dyevert_not_used_inj += 1
                            vol_not_used_inj += row[CONTRAST_VOLUME_TO_PATIENT]
                        elif row[IS_AN_INJECTION] == 1 and puff_inj == 1:
                            dyevert_used_inj += 1
                            vol_used_inj += row[CONTRAST_VOLUME_TO_PATIENT]
                        elif row[IS_AN_INJECTION] == 1 and row[PERCENT_CONTRAST_SAVED] == 0 and puff_inj == 2:
                            dyevert_not_used_puff += 1
                            vol_not_used_puff += row[CONTRAST_VOLUME_TO_PATIENT]
                        elif row[IS_AN_INJECTION] == 1 and puff_inj == 2:
                            dyevert_used_puff += 1
                            vol_used_puff += row[CONTRAST_VOLUME_TO_PATIENT]

    uses.append([dyevert_not_used_inj, dyevert_used_inj, dyevert_not_used_puff, dyevert_used_puff, first_inj, last_inj])
    while len(uses) <= number_of_cases+2:
        uses.append([0, 0, 0, 0, "", ""])
    uses = uses[1:]
    return uses


def case_reconstruction(caseID, cmsw):
    return []


def sort_by_date(case):
    return case[0], case[1]


def case_by_date(filenames):
    cases = []
    for filename in filenames:
        con = sqlite.connect(filename)
        with con:
            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWCases')
            rows = cur.fetchall()
            for row in rows:
                cases.append(row[5][0:7])
    cases.sort()
    casesPerMonth = {}
    for case in cases:
        if not casesPerMonth.get(case) is None:
            casesPerMonth[case] += 1
        else:
            casesPerMonth[case] = 1
    CasesPerQuarter = {}
    for month in casesPerMonth:
        if CasesPerQuarter.get(str(month[0:-3])+' Q'+str(math.ceil(float(month[-2:])/3))) is None:
            CasesPerQuarter[str(month[0:-3])+' Q'+str(math.ceil(float(month[-2:])/3))] = int(casesPerMonth.get(month))
        else:
            CasesPerQuarter[str(month[0:-3])+' Q'+str(math.ceil(float(month[-2:])/3))] += int(casesPerMonth.get(month))
    return [casesPerMonth, CasesPerQuarter]


def excel_flag_write(file_names, cmsws):
    """Writes data into the an Excel Sheet
    Takes two inputs:
        -the file names of the CMSW databases
        -the serial numbers of the CMSWs
    Generates one files:
        -The flagged table, with data that hits possible removal criteria being highlighted in yellow
    """
    print('Processing DyeMinish data with flagging')
    cases = list_builder(file_names)
    dvuses = dyevert_uses(file_names)
    xlsx_name = str(cmsws) + 'DyeMinishFlaggedOutput.xlsx'
    wb = openpyxl.load_workbook('Dyeminish-template.xlsx')
    data_sheet = wb.active
    data_sheet.title = 'Sheet1'
    print('Writing DyeMinish data')
    for file_name in file_names:
        con = sqlite.connect(file_name)
        with con:
            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWCases')
            rows = cur.fetchall()
        global TOTAL_DURATION, END_TIME, IS_AN_INJECTION, LINEAR_DYEVERT_MOVEMENT, DYEVERT_CONTRAST_VOLUME_DIVERTED, \
            DYEVERT_CONTRAST_VOLUME_DIVERTED, PERCENT_CONTRAST_SAVED, CONTRAST_VOLUME_TO_PATIENT, \
            PREDOMINANT_CONTRAST_LINE_PRESSURE, FLOW_RATE_TO_FROM_SYRINGE, FLOW_RATE_TO_PATIENT
        missing_summary = []
        if not rows == []:
            if rows[0][2] == '2.1.56' or rows[0][2] == '2.1.24' or rows[0][2] == '2.1.67':
                TOTAL_DURATION = 19
                END_TIME = 20
                IS_AN_INJECTION = 5
                LINEAR_DYEVERT_MOVEMENT = 15
                DYEVERT_CONTRAST_VOLUME_DIVERTED = 17
                PERCENT_CONTRAST_SAVED = 18
                CONTRAST_VOLUME_TO_PATIENT = 20
                FLOW_RATE_TO_FROM_SYRINGE = 28
                FLOW_RATE_TO_PATIENT = 29
                PREDOMINANT_CONTRAST_LINE_PRESSURE = 30
            else:
                TOTAL_DURATION = 20
                END_TIME = 19
                IS_AN_INJECTION = 8
                LINEAR_DYEVERT_MOVEMENT = 18
                DYEVERT_CONTRAST_VOLUME_DIVERTED = 20
                PERCENT_CONTRAST_SAVED = 21
                CONTRAST_VOLUME_TO_PATIENT = 23
                FLOW_RATE_TO_FROM_SYRINGE = 32
                FLOW_RATE_TO_PATIENT = 33
                PREDOMINANT_CONTRAST_LINE_PRESSURE = 34
            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWCases')
            rows = cur.fetchall()

            for row in rows:
                if row[ATTEMPTED_CONTRAST_INJECTION_VOLUME] == row[DIVERTED_CONTRAST_VOLUME] == \
                        row[CUMULATIVE_VOLUME_TO_PATIENT] == row[DIVERTED_CONTRAST_VOLUME] == 0:
                    missing_summary.append([row[0], file_name])
                    logging.warning('Case ' + str(row[0]) + ' may be missing summary data')
        with con:

            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWInjections')
            rows = cur.fetchall()
            has_injections = []
            for row in rows:
                has_injections.append([row[1], file_name])
    rg17err = []
    for missing in missing_summary:
        if missing in has_injections:
            rg17err.append(missing)
    if not rg17err == []:
        for entry in range(len(rg17err)):
            con = sqlite.connect(rg17err[entry][1])
            with con:
                cur = con.cursor()
                cur.execute('SELECT * FROM CMSWCases')
                rows = cur.fetchall()
            rg17err[entry] = [rg17err[entry][0], rg17err[entry][1], rows[rg17err[entry][0]+1][1]]
    for row in range(len(cases)):
        for col in range(40):
            if cases[row][13] == 'PM':
                data_sheet.cell(row=row + 2, column=col + 1).fill = \
                    PatternFill(fill_type='solid', start_color=YELLOW, end_color=YELLOW)
                print("PM")
            elif float(cases[row][2]) <= 5.:
                data_sheet.cell(row=row + 2, column=col + 1).fill = \
                    PatternFill(fill_type='solid', start_color=YELLOW, end_color=YELLOW)
                print("Duration < 5 min")
            elif cases[row][4] == 0 and cases[row][5] == 0 and cases[row][6] == 0 and cases[row][7] == 0:
                data_sheet.cell(row=row + 2, column=col + 1).fill = PatternFill(
                    fill_type='solid', start_color=YELLOW, end_color=YELLOW)
                print("No DyeVert use")
            elif cases[row][5] <= 5:
                data_sheet.cell(row=row + 2, column=col + 1).fill = PatternFill(
                    fill_type='solid', start_color=YELLOW, end_color=YELLOW)
            data_sheet.cell(row=row + 2, column=col + 1).alignment = Alignment(wrapText=True)
        data_sheet.cell(row=row+2, column=4, value=cases[row][0])
        data_sheet.cell(row=row+2, column=5, value=cases[row][1])
        data_sheet.cell(row=row+2, column=6, value=cases[row][15])
        data_sheet.cell(row=row+2, column=7, value=cases[row][2])
        data_sheet.cell(row=row+2, column=8, value=cases[row][3])
        data_sheet.cell(row=row+2, column=9, value=cases[row][4])
        data_sheet.cell(row=row+2, column=10, value=cases[row][5])
        data_sheet.cell(row=row+2, column=11, value=cases[row][6])
        data_sheet.cell(row=row+2, column=12, value=cases[row][7])
        data_sheet.cell(row=row+2, column=13, value=cases[row][8])
        data_sheet.cell(row=row+2, column=19, value=cases[row][9])
        data_sheet.cell(row=row+2, column=20, value=cases[row][12])
        data_sheet.cell(row=row+2, column=22, value=cases[row][10])
        data_sheet.cell(row=row+2, column=23, value=cases[row][11])
        if cases[row][13] == 'PM':
            data_sheet.cell(row=row + 2, column=16, value='DyeTect Case')
            data_sheet.cell(row=row + 2, column=14, value='No')
        elif float(cases[row][2]) <= 5.:
            data_sheet.cell(row=row + 2, column=16, value='Case less than 5 Minutes')
            data_sheet.cell(row=row + 2, column=14, value='No')
        elif cases[row][4] == 0 and cases[row][5] == 0 and cases[row][6] == 0 and cases[row][7] == 0:
            errset = 0
            for err in rg17err:
                if cases[row][14] == err[0]:
                    data_sheet.cell(row=row + 2, column=16, value='RG17 Error')
                    errset = 1
            if errset == 0:
                data_sheet.cell(row=row + 2, column=16, value='No contrast was injected')
            data_sheet.cell(row=row + 2, column=14, value='No')

        elif cases[row][4] == 0 and cases[row][5] == 0 and cases[row][6] == 0 and cases[row][7] == 0 :
            data_sheet.cell(row=row + 2, column=16, value='No contrast was injected')
            data_sheet.cell(row=row + 2, column=14, value='No')
        # write the case type (pmdv) into column 36
        data_sheet.cell(row=row+2, column=30, value=dvuses[row][1])
        data_sheet.cell(row=row+2, column=31, value=dvuses[row][0])
        data_sheet.cell(row=row+2, column=32, value=dvuses[row][3])
        data_sheet.cell(row=row+2, column=33, value=dvuses[row][2])
        data_sheet.cell(row=row+2, column=36, value=cases[row][13])
        data_sheet.cell(row=row+2, column=38, value=dvuses[row][4])
        data_sheet.cell(row=row+2, column=39, value=dvuses[row][5])

    date_sheet = wb.create_sheet(title='Cases by Date')
    dates = case_by_date(file_names)
    months = ['Month']
    month_case_count = ['Cases in Month']
    for date in list(dates[0]):
        months.append(date)
        month_case_count.append(dates[0][date])
    quarters = ['Quarter']
    quarter_cases = ['Cases Per Quarter']
    for Q in list(dates[1]):
        quarters.append(Q)
        quarter_cases.append(dates[1][Q])
    col = 1
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for month in months:
        date_sheet.cell(row=1, column=col, value=month)
        date_sheet.cell(row=1, column=col).alignment = Alignment(wrapText=True)
        date_sheet.cell(row=1, column=col).border = thin_border
        col += 1
    col = 1
    for count in month_case_count:
        date_sheet.cell(row=2, column=col, value=count)
        date_sheet.cell(row=2, column=col).alignment = Alignment(wrapText=True)
        date_sheet.cell(row=2, column=col).border = thin_border
        col += 1
    col = 1
    for quarter in quarters:
        date_sheet.cell(row=4, column=col, value=quarter)
        date_sheet.cell(row=4, column=col).alignment = Alignment(wrapText=True)
        date_sheet.cell(row=4, column=col).border = thin_border
        col += 1
    col = 1
    for count in quarter_cases:
        date_sheet.cell(row=5, column=col, value=count)
        date_sheet.cell(row=5, column=col).alignment = Alignment(wrapText=True)
        date_sheet.cell(row=5, column=col).border = thin_border
        col += 1
    c1 = LineChart()
    c1.title = 'Number of DyeVert Cases by Month'
    c1.y_axis.title = '# of cases'
    c1.x_axis.title = ''
    c1.style = 13

    data = Reference(date_sheet, 2, 1, len(months), 2)
    c1.add_data(data)

    s1 = c1.series[0]
    s1.marker.symbol = 'square'

    date_sheet.add_chart(c1, 'A7')
    wb.save(xlsx_name)
    print('DyeMinish report with flagged data finished')


def excel_destructive_write(file_names, cmsws):
    """Writes data into the an Excel Sheet
    Takes two inputs:
        -the file names of the CMSW databases
        -the serial numbers of the CMSWs
    Generates one files:
        -The flagged table, with data that hits possible removal criteria being excluded
    """
    print('Processing DyeMinish data with deletion')
    cases = list_builder(file_names)
    remove_cases = []
    print('Removing cases')
    for case in cases[0]:
        if float(case[6]) <= 5. or int(case[8]) == int(case[9]) == int(case[10]) == int(case[11]) == 0 or \
                case[20] == 0 or case[20] == 'PM':
            remove_cases.append(case)

    for case in remove_cases:
        cases[0].remove(case)

    xlsx_name = str(cmsws) + 'DyeMinishFilteredOutput.xlsx'
    wb = openpyxl.load_workbook('Dyeminish-template.xlsx')
    data_sheet = wb.active
    data_sheet.title = 'Sheet1'
    print('Writing cleaned DyeMinish data')
    for row in range(len(cases[0])):
        for col in range(len(cases[0][row]) - 1):
            data_sheet.cell(row=row + 2, column=col + 1, value=cases[0][row][col])
            data_sheet.cell(row=row + 2, column=col + 1).alignment = Alignment(wrapText=True)

    wb.save(xlsx_name)
    print('DyeMinish report with deletion finished')

import sqlite3 as sqlite
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
import cmsw_read

# case column numbers
CMSW_CASE_ID = 0
CASE_ID = 1
SERIAL_NUMBER = 3
DATE_OF_PROCEDURE = 5
THRESHOLD_VOLUME = 8
ATTEMPTED_CONTRAST_INJECTION_VOLUME = 13
DIVERTED_CONTRAST_VOLUME = 14
CUMULATIVE_VOLUME_TO_PATIENT = 15
PERCENTAGE_CONTRAST_DIVERTED = 16
# Total duration.  
# 19 for CMSW 
# 20 for iPad 
TOTAL_DURATION = 20

#
# Use these for iPad 
#  injection column numbers
TIME_STAMP = 2
SYRINGE_REVISION = 3
PMDV_REVISION = 4
IS_AN_INJECTION = 8
IS_ASPIRATING_CONTRAST = 9
DYEVERT_DIAMETER = 10
SYRINGE_DIAMETER = 11
STARTING_SYRINGE_POSITION = 12
ENDING_SYRINGE_POSITION = 13
LINEAR_SYRINGE_MOVEMENT = 14
SYRINGE_VOLUME_INJECTED_OR_ASPIRATED = 15
STARTING_DYEVERT_POSITION = 16
ENDING_DYEVERT_POSITION = 17
LINEAR_DYEVERT_MOVEMENT = 18
DIVERT_VOLUME_DIVERTED = 19
DYEVERT_CONTRAST_VOLUME_DIVERTED = 20
PERCENT_CONTRAST_SAVED = 21
INJECTION_VOLUME_TO_PATIENT = 22
CONTRAST_VOLUME_TO_PATIENT = 23
CUMULATIVE_CONTRAST_VOLUME_TO_PATIENT = 24
OTHER_VOLUME_TO_PATIENT = 25
STARTING_CONTRAST_PERCENT_IN_SYRINGE = 27
STARTING_CONTRAST_PERCENT_IN_DYEVERT = 29
ENDING_CONTRAST_PERCENT_IN_DYEVERT = 30
DURATION = 31
FLOW_RATE_TO_FROM_SYRINGE = 32
FLOW_RATE_TO_PATIENT = 33
PREDOMINANT_CONTRAST_LINE_PRESSURE = 34
STARTING_DYEVERT_STOPCOCK_POSITION = 35
IS_SYSTEM_PAUSED = 36
ENDING_CONTRAST_PERCENT_IN_SYRINGE = 28
SYRINGE_ADDRESS = 5
PMDV_ADDRESS = 4
IS_DEVICE_REPLACEMENT = 7

#
# Use these for CMSW 
#  injection column numbers
# TIME_STAMP = 2
# SYRINGE_REVISION = 3
# PMDV_REVISION = 4
# IS_AN_INJECTION = 5
# IS_ASPIRATING_CONTRAST = 6
# DYEVERT_DIAMETER = 7
# SYRINGE_DIAMETER = 8
# STARTING_SYRINGE_POSITION = 9
# ENDING_SYRINGE_POSITION = 10
# LINEAR_SYRINGE_MOVEMENT = 11
# SYRINGE_VOLUME_INJECTED_OR_ASPIRATED = 12
# STARTING_DYEVERT_POSITION = 13
# ENDING_DYEVERT_POSITION = 14
# LINEAR_DYEVERT_MOVEMENT = 15
# DIVERT_VOLUME_DIVERTED = 16
# DYEVERT_CONTRAST_VOLUME_DIVERTED = 17
# PERCENT_CONTRAST_SAVED = 18
# INJECTION_VOLUME_TO_PATIENT = 19
# CONTRAST_VOLUME_TO_PATIENT = 20
# CUMULATIVE_CONTRAST_VOLUME_TO_PATIENT = 21
# OTHER_VOLUME_TO_PATIENT = 22
# STARTING_CONTRAST_PERCENT_IN_SYRINGE = 24
# STARTING_CONTRAST_PERCENT_IN_DYEVERT = 25
# ENDING_CONTRAST_PERCENT_IN_DYEVERT = 26
# DURATION = 27
# FLOW_RATE_TO_FROM_SYRINGE = 28
# FLOW_RATE_TO_PATIENT = 29
# PREDOMINANT_CONTRAST_LINE_PRESSURE = 30
# STARTING_DYEVERT_STOPCOCK_POSITION = 31
# IS_SYSTEM_PAUSED = 32
# ENDING_CONTRAST_PERCENT_IN_SYRINGE = 33
# SYRINGE_ADDRESS = 34
# PMDV_ADDRESS = 35
# IS_DEVICE_REPLACEMENT = 36


# colors
WHITE = 0
LTGRN = 1
GREEN = 2
YELLOW = 3
RED = 4

# RGB codes
BLUE = '2F6BC0'
YLW = 'FBE798'
GRN = 'C5E1B3'


# Write data for sales team to appropriate template


def _sort_criteria(case):
    """reads info for a sort"""
    return case[1], case[2]


def injection_table(file_names, cmsw):
    """Connects to an individual database and determines which injections were puffs, injections,
    and leaves some uncatagorized to be classified by a person looking at the surrounding data
    """
    xlsx2_name = str(cmsw).replace('s', '') + 'rods-detailed-data.xlsx'
    wb = Workbook(write_only=True)
    data_sheet = wb.create_sheet()
    yellow = PatternFill(fill_type="solid", start_color=YLW, end_color=YLW)
    green = PatternFill(fill_type="solid", start_color=GRN, end_color=GRN)
    alignment = Alignment(wrap_text=True, shrink_to_fit=True)
    data_sheet.title = 'Sheet1'
    data_sheet.row_dimensions.group(2, 3, hidden=True)
    data_sheet.column_dimensions.group('B', 'E', hidden=True)
    data_sheet.column_dimensions.group('H', 'R', hidden=True)
    data_sheet.column_dimensions.group('U', 'U', hidden=True)
    data_sheet.column_dimensions.group('X', 'AC', hidden=True)
    data_sheet.column_dimensions.group('AF', 'AI', hidden=True)
    # values for width do not match actual values
    # columns ending in .42 are .71 over the actual width
    # columns ending in .01 are .72 over the actual width
    data_sheet.column_dimensions['A'].width = 18.42
    data_sheet.column_dimensions['G'].width = 10.01
    data_sheet.column_dimensions['V'].width = 12.01
    data_sheet.column_dimensions['W'].width = 11.01
    data_sheet.column_dimensions['AD'].width = 14.42
    data_sheet.column_dimensions['AE'].width = 10.42
    data_sheet.column_dimensions['AJ'].width = 12.42
    data_sheet.freeze_panes = 'A2'
    frmcell = [WriteOnlyCell(ws=data_sheet, value='Contrast Aspirated'),
               WriteOnlyCell(ws=data_sheet, value='Contrast Diverted'),
               WriteOnlyCell(ws=data_sheet, value='Percent Saved'),
               WriteOnlyCell(ws=data_sheet, value='Contrast to Patient'),
               WriteOnlyCell(ws=data_sheet, value='Cumulative'),
               WriteOnlyCell(ws=data_sheet, value='Flow Rate from Syringe'),
               WriteOnlyCell(ws=data_sheet, value='Flow Rate to Patient'),
               WriteOnlyCell(ws=data_sheet, value='Volume Attempted')
               ]
    empty_cell = WriteOnlyCell(ws=data_sheet, value='')
    for cell in range(len(frmcell)):
        frmcell[cell].font = Font(bold=True)
    cases = [[empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, frmcell[0], empty_cell,
              empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell,
              empty_cell, empty_cell, frmcell[1], frmcell[2], empty_cell, frmcell[3], frmcell[4], empty_cell,
              empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, frmcell[5], frmcell[6], empty_cell,
              empty_cell, empty_cell, empty_cell, frmcell[7]],
             [empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell,
              empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell,
              empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell,
              empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell,
              empty_cell, empty_cell, empty_cell, empty_cell],
             [WriteOnlyCell(ws=data_sheet, value='Time Stamp'), WriteOnlyCell(ws=data_sheet, value='Syringe Revision'),
              WriteOnlyCell(ws=data_sheet, value='PMDV Revision'),
              WriteOnlyCell(ws=data_sheet, value='Syringe Address'),
              WriteOnlyCell(ws=data_sheet, value='PMDV Address'),
              WriteOnlyCell(ws=data_sheet, value='Injection or Aspiration'),
              WriteOnlyCell(ws=data_sheet, value='Aspirating Contrast'),
              WriteOnlyCell(ws=data_sheet, value='Replacing Device'),
              WriteOnlyCell(ws=data_sheet, value='Current DyeVert Diameter'),
              WriteOnlyCell(ws=data_sheet, value='Current Syringe Diameter'),
              WriteOnlyCell(ws=data_sheet, value='Starting Syringe Plunger Position'),
              WriteOnlyCell(ws=data_sheet, value='Ending Syringe Plunger Position'),
              WriteOnlyCell(ws=data_sheet, value='Syringe Linear Plunger Position'),
              WriteOnlyCell(ws=data_sheet, value='Volume(Injected / Aspirated)'),
              WriteOnlyCell(ws=data_sheet, value='Starting DyeVert Plus Reservoir Plunger Position'),
              WriteOnlyCell(ws=data_sheet, value='Ending DyeVert Plus Reservoir Plunger Position'),
              WriteOnlyCell(ws=data_sheet, value='DyeVert Plus Reservoir Linear Plunger Position'),
              WriteOnlyCell(ws=data_sheet, value='DyeVert Plus Reservoir Volume Diverted'),
              WriteOnlyCell(ws=data_sheet, value='DyeVert Plus Reservoir Contrast Volume Diverted'),
              WriteOnlyCell(ws=data_sheet, value='PercentContrastSaved'),
              WriteOnlyCell(ws=data_sheet, value='Total Injection Volume to Patient'),
              WriteOnlyCell(ws=data_sheet, value='Volume of Contrast Injected'),
              WriteOnlyCell(ws=data_sheet, value='Cumulative Contrast Volume Injected'),
              WriteOnlyCell(ws=data_sheet, value='Volume of Other Injected'),
              WriteOnlyCell(ws=data_sheet, value='Starting Contrast Percentage in Syringe'),
              WriteOnlyCell(ws=data_sheet, value='Ending Contrast Percentage in Syringe'),
              WriteOnlyCell(ws=data_sheet, value='Starting Contrast Percentage in DyeVert Plus Reservoir'),
              WriteOnlyCell(ws=data_sheet, value='Ending Contrast Percentage in DyeVert Plus Reservoir'),
              WriteOnlyCell(ws=data_sheet, value='Duration'),
              WriteOnlyCell(ws=data_sheet, value='Flow Rate from Syringe'),
              WriteOnlyCell(ws=data_sheet, value='Flow Rate to Patient'),
              WriteOnlyCell(ws=data_sheet, value='Contrast Line Pressure'),
              WriteOnlyCell(ws=data_sheet, value='DyeVert Plus Stopcock Position'),
              WriteOnlyCell(ws=data_sheet, value='System IsSystemPaused'), empty_cell, empty_cell]
             ]
    case_number = 0

    for file_name in file_names:

        con = sqlite.connect(file_name)

        with con:

            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWCases')
            rows = cur.fetchall()
            case_id_number = {}
            global TIME_STAMP, SYRINGE_REVISION, PMDV_REVISION, IS_AN_INJECTION, IS_ASPIRATING_CONTRAST, \
                DYEVERT_DIAMETER, SYRINGE_DIAMETER, STARTING_SYRINGE_POSITION, ENDING_SYRINGE_POSITION, \
                LINEAR_SYRINGE_MOVEMENT, SYRINGE_VOLUME_INJECTED_OR_ASPIRATED, STARTING_DYEVERT_POSITION, \
                ENDING_DYEVERT_POSITION, LINEAR_DYEVERT_MOVEMENT, DIVERT_VOLUME_DIVERTED, \
                DYEVERT_CONTRAST_VOLUME_DIVERTED, PERCENT_CONTRAST_SAVED, INJECTION_VOLUME_TO_PATIENT, \
                CONTRAST_VOLUME_TO_PATIENT, CUMULATIVE_CONTRAST_VOLUME_TO_PATIENT, OTHER_VOLUME_TO_PATIENT, \
                STARTING_CONTRAST_PERCENT_IN_SYRINGE, STARTING_CONTRAST_PERCENT_IN_DYEVERT, \
                ENDING_CONTRAST_PERCENT_IN_DYEVERT, DURATION, FLOW_RATE_TO_FROM_SYRINGE, FLOW_RATE_TO_PATIENT, \
                PREDOMINANT_CONTRAST_LINE_PRESSURE, STARTING_DYEVERT_STOPCOCK_POSITION, IS_SYSTEM_PAUSED, \
                ENDING_CONTRAST_PERCENT_IN_SYRINGE, SYRINGE_ADDRESS, PMDV_ADDRESS, IS_DEVICE_REPLACEMENT, \
                SERIAL_NUMBER, DATE_OF_PROCEDURE, THRESHOLD_VOLUME, ATTEMPTED_CONTRAST_INJECTION_VOLUME, \
                DIVERTED_CONTRAST_VOLUME, CUMULATIVE_VOLUME_TO_PATIENT, PERCENTAGE_CONTRAST_DIVERTED, TOTAL_DURATION
            if not rows == []:
                if rows[0][2] == '2.1.56' or rows[0][2] == '2.1.24' or rows[0][2] == '2.1.67' or \
                        rows[0][2] == '2.0.1981' or rows[0][2] == '2.0.2013':
                    TIME_STAMP = 2
                    SYRINGE_REVISION = 3
                    PMDV_REVISION = 4
                    IS_AN_INJECTION = 5
                    IS_ASPIRATING_CONTRAST = 6
                    DYEVERT_DIAMETER = 7
                    SYRINGE_DIAMETER = 8
                    STARTING_SYRINGE_POSITION = 9
                    ENDING_SYRINGE_POSITION = 10
                    LINEAR_SYRINGE_MOVEMENT = 11
                    SYRINGE_VOLUME_INJECTED_OR_ASPIRATED = 12
                    STARTING_DYEVERT_POSITION = 13
                    ENDING_DYEVERT_POSITION = 14
                    LINEAR_DYEVERT_MOVEMENT = 15
                    DIVERT_VOLUME_DIVERTED = 16
                    DYEVERT_CONTRAST_VOLUME_DIVERTED = 17
                    PERCENT_CONTRAST_SAVED = 18
                    INJECTION_VOLUME_TO_PATIENT = 19
                    CONTRAST_VOLUME_TO_PATIENT = 20
                    CUMULATIVE_CONTRAST_VOLUME_TO_PATIENT = 21
                    OTHER_VOLUME_TO_PATIENT = 22
                    STARTING_CONTRAST_PERCENT_IN_SYRINGE = 24
                    STARTING_CONTRAST_PERCENT_IN_DYEVERT = 25
                    ENDING_CONTRAST_PERCENT_IN_DYEVERT = 26
                    DURATION = 27
                    FLOW_RATE_TO_FROM_SYRINGE = 28
                    FLOW_RATE_TO_PATIENT = 29
                    PREDOMINANT_CONTRAST_LINE_PRESSURE = 30
                    STARTING_DYEVERT_STOPCOCK_POSITION = 31
                    IS_SYSTEM_PAUSED = 32
                    ENDING_CONTRAST_PERCENT_IN_SYRINGE = 33
                    SYRINGE_ADDRESS = 34
                    PMDV_ADDRESS = 35
                    IS_DEVICE_REPLACEMENT = 36
                else:
                    SERIAL_NUMBER = 3
                    DATE_OF_PROCEDURE = 5
                    THRESHOLD_VOLUME = 8
                    ATTEMPTED_CONTRAST_INJECTION_VOLUME = 13
                    DIVERTED_CONTRAST_VOLUME = 14
                    CUMULATIVE_VOLUME_TO_PATIENT = 15
                    PERCENTAGE_CONTRAST_DIVERTED = 16
                    TOTAL_DURATION = 20
                    TIME_STAMP = 2
                    SYRINGE_REVISION = 3
                    PMDV_REVISION = 4
                    IS_AN_INJECTION = 8
                    IS_ASPIRATING_CONTRAST = 9
                    DYEVERT_DIAMETER = 10
                    SYRINGE_DIAMETER = 11
                    STARTING_SYRINGE_POSITION = 12
                    ENDING_SYRINGE_POSITION = 13
                    LINEAR_SYRINGE_MOVEMENT = 14
                    SYRINGE_VOLUME_INJECTED_OR_ASPIRATED = 15
                    STARTING_DYEVERT_POSITION = 16
                    ENDING_DYEVERT_POSITION = 17
                    LINEAR_DYEVERT_MOVEMENT = 18
                    DIVERT_VOLUME_DIVERTED = 19
                    DYEVERT_CONTRAST_VOLUME_DIVERTED = 20
                    PERCENT_CONTRAST_SAVED = 21
                    INJECTION_VOLUME_TO_PATIENT = 22
                    CONTRAST_VOLUME_TO_PATIENT = 23
                    CUMULATIVE_CONTRAST_VOLUME_TO_PATIENT = 24
                    OTHER_VOLUME_TO_PATIENT = 25
                    STARTING_CONTRAST_PERCENT_IN_SYRINGE = 27
                    STARTING_CONTRAST_PERCENT_IN_DYEVERT = 29
                    ENDING_CONTRAST_PERCENT_IN_DYEVERT = 30
                    DURATION = 31
                    FLOW_RATE_TO_FROM_SYRINGE = 32
                    FLOW_RATE_TO_PATIENT = 33
                    PREDOMINANT_CONTRAST_LINE_PRESSURE = 34
                    STARTING_DYEVERT_STOPCOCK_POSITION = 35
                    IS_SYSTEM_PAUSED = 36
                    ENDING_CONTRAST_PERCENT_IN_SYRINGE = 28
                    SYRINGE_ADDRESS = 5
                    PMDV_ADDRESS = 4
                    IS_DEVICE_REPLACEMENT = 7
            for row in rows:
                case_id_number[row[CMSW_CASE_ID]] = row[CASE_ID][-23:-4]

        with con:

            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWInjections')
            rows = cur.fetchall()

            for row in rows:
                perc_saved = WriteOnlyCell(value=row[PERCENT_CONTRAST_SAVED], ws=data_sheet)
                if case_number != row[CASE_ID]:
                    case_number = row[CASE_ID]
                    _cmsw = cmsw_read.cmsw_id_read(file_name)
                    cases.append([WriteOnlyCell(ws=data_sheet, value='CMSW'), empty_cell, empty_cell, empty_cell,
                                  empty_cell, WriteOnlyCell(ws=data_sheet, value=_cmsw), empty_cell, empty_cell,
                                  empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell,
                                  empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell])
                    cases.append([WriteOnlyCell(ws=data_sheet, value='Case'), empty_cell, empty_cell, empty_cell,
                                  empty_cell, WriteOnlyCell(ws=data_sheet, value=case_id_number[case_number]),
                                  empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell,
                                  empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell, empty_cell,
                                  empty_cell])
                if row[IS_AN_INJECTION] == 1:
                    inj_asp = 'INJ'
                else:
                    inj_asp = 'ASP'
                if row[IS_ASPIRATING_CONTRAST] == 1:
                    contrast_asp = 'Yes'
                else:
                    contrast_asp = ''
                if row[IS_DEVICE_REPLACEMENT] == 1:
                    replacement = 'Yes'
                else:
                    replacement = ''
                if inj_asp == 'INJ':
                    perc_saved.font = Font(bold=True)
                    if row[CONTRAST_VOLUME_TO_PATIENT] + row[DYEVERT_CONTRAST_VOLUME_DIVERTED] >= 3:
                        puff_inj = WriteOnlyCell(value='Injection', ws=data_sheet)
                        puff_inj.font = Font(bold=True)
                    elif row[CONTRAST_VOLUME_TO_PATIENT] + row[DYEVERT_CONTRAST_VOLUME_DIVERTED] <= 2:
                        puff_inj = WriteOnlyCell(value='Puff', ws=data_sheet)
                        puff_inj.font = Font(bold=True)
                    elif row[FLOW_RATE_TO_FROM_SYRINGE] >= 2.5:
                        puff_inj = WriteOnlyCell(value='Injection', ws=data_sheet)
                        puff_inj.font = Font(bold=True)
                    elif row[FLOW_RATE_TO_FROM_SYRINGE] <= 2:
                        puff_inj = WriteOnlyCell(value='Puff', ws=data_sheet)
                        puff_inj.font = Font(bold=True)
                    else:
                        debug_msg = 'Event ' + str(row[0]) + ' in cmsw ' + str(cmsw_read.cmsw_id_read(file_name)) + \
                                    ', case ' + str(row[CASE_ID]) + ' matched neither type'
                        logging.warning(debug_msg)
                        print(debug_msg)
                        puff_inj = empty_cell
                else:
                    puff_inj = empty_cell
                    perc_saved.font = Font(bold=False)
                    for entry in range(len(row)):
                        if row[entry] is None:
                            newrow = list(row)
                            newrow[entry] = '0'
                            row = newrow
                cases.append([WriteOnlyCell(ws=data_sheet, value=row[TIME_STAMP]),
                              WriteOnlyCell(ws=data_sheet, value=row[SYRINGE_REVISION]),
                              WriteOnlyCell(ws=data_sheet, value=row[PMDV_REVISION]),
                              WriteOnlyCell(ws=data_sheet, value=row[SYRINGE_ADDRESS]),
                              WriteOnlyCell(ws=data_sheet, value=row[PMDV_ADDRESS]),
                              WriteOnlyCell(ws=data_sheet, value=inj_asp),
                              WriteOnlyCell(ws=data_sheet, value=contrast_asp),
                              WriteOnlyCell(ws=data_sheet, value=replacement),
                              WriteOnlyCell(ws=data_sheet, value=row[DYEVERT_DIAMETER]),
                              WriteOnlyCell(ws=data_sheet, value=row[SYRINGE_DIAMETER]),
                              WriteOnlyCell(ws=data_sheet, value=row[STARTING_SYRINGE_POSITION]),
                              WriteOnlyCell(ws=data_sheet, value=row[ENDING_SYRINGE_POSITION]),
                              WriteOnlyCell(ws=data_sheet, value=row[LINEAR_SYRINGE_MOVEMENT]),
                              WriteOnlyCell(ws=data_sheet, value=row[SYRINGE_VOLUME_INJECTED_OR_ASPIRATED]),
                              WriteOnlyCell(ws=data_sheet, value=row[STARTING_DYEVERT_POSITION]),
                              WriteOnlyCell(ws=data_sheet, value=row[ENDING_DYEVERT_POSITION]),
                              WriteOnlyCell(ws=data_sheet, value=row[LINEAR_DYEVERT_MOVEMENT]),
                              WriteOnlyCell(ws=data_sheet, value=row[DIVERT_VOLUME_DIVERTED]),
                              WriteOnlyCell(ws=data_sheet, value=round(row[DYEVERT_CONTRAST_VOLUME_DIVERTED], 2)),
                              perc_saved, WriteOnlyCell(ws=data_sheet, value=row[INJECTION_VOLUME_TO_PATIENT]),
                              WriteOnlyCell(ws=data_sheet, value=round(row[CONTRAST_VOLUME_TO_PATIENT], 2)),
                              WriteOnlyCell(ws=data_sheet, value=round(row[CUMULATIVE_CONTRAST_VOLUME_TO_PATIENT], 2)),
                              WriteOnlyCell(ws=data_sheet, value=row[OTHER_VOLUME_TO_PATIENT]),
                              WriteOnlyCell(ws=data_sheet, value=row[STARTING_CONTRAST_PERCENT_IN_SYRINGE]),
                              WriteOnlyCell(ws=data_sheet, value=row[ENDING_CONTRAST_PERCENT_IN_SYRINGE]),
                              WriteOnlyCell(ws=data_sheet, value=row[STARTING_CONTRAST_PERCENT_IN_DYEVERT]),
                              WriteOnlyCell(ws=data_sheet, value=row[ENDING_CONTRAST_PERCENT_IN_DYEVERT]),
                              WriteOnlyCell(ws=data_sheet, value=row[DURATION]),
                              WriteOnlyCell(ws=data_sheet, value=round(float(row[FLOW_RATE_TO_FROM_SYRINGE]), 2)),
                              WriteOnlyCell(ws=data_sheet, value=round(float(row[FLOW_RATE_TO_PATIENT]), 2)),
                              WriteOnlyCell(ws=data_sheet, value=row[PREDOMINANT_CONTRAST_LINE_PRESSURE]),
                              WriteOnlyCell(ws=data_sheet, value=row[STARTING_DYEVERT_STOPCOCK_POSITION]),
                              WriteOnlyCell(ws=data_sheet, value=row[IS_SYSTEM_PAUSED]),
                              WriteOnlyCell(ws=data_sheet, value=''),
                              WriteOnlyCell(ws=data_sheet, value=round(float(row[CONTRAST_VOLUME_TO_PATIENT])
                                                                    + float(row[DYEVERT_CONTRAST_VOLUME_DIVERTED]), 2)),
                              puff_inj])
                cases[-1][29].font = Font(color=BLUE)
                cases[-1][30].font = Font(color=BLUE)

    print('Applying formatting', end='')
    for case in range(len(cases)):
        if case % 10000 == 0:
            print('.', end='')
        for cell in range(len(cases[case])):
            _cell = cases[case][cell]
            _cell.alignment = alignment
            if cases[case][5].internal_value == 'INJ' and _cell != empty_cell:
                if cases[case][19].internal_value == 0:
                    _cell.fill = yellow
                if case > 3 and cases[case][19].internal_value > 0:
                    _cell.fill = green
            if (cases[case][5].internal_value == 'ASP' and cases[case][6].internal_value != 'Yes') or \
                    (cases[case][5].internal_value == 'INJ' and (int(cases[case][29].internal_value) == 0
                                                                 or int(cases[case][21].internal_value) == 0)):
                data_sheet.row_dimensions[case + 1].hidden = True
            cases[case][cell] = _cell
    print('')
    print('Writing injection data', end='')
    for case in range(len(cases)):
        # print('Writing event ',case, 'of ',len(cases))
        if case % 100 == 0:
            print('.', end='')
            print('Writing event ', case, 'of ', len(cases))
        data_sheet.append(cases[case])
    print('')
    print('Saving...')
    wb.save(xlsx2_name)


def list_builder(file_names):
    """Takes the list of files and builds the list of lists to write"""
    cases = []

    for file_name in file_names:
        con = sqlite.connect(file_name)
        with con:
            cur = con.cursor()
            cur.execute('SELECT * FROM CMSWCases')
            rows = cur.fetchall()
            uses = dyevert_uses(file_name)
            for row in rows:
                # don't do the exclusion anymore. Updating this file to no longer reject excluded cases.
                # if not (row[TOTAL_DURATION] <= 5) and not (row[ATTEMPTED_CONTRAST_INJECTION_VOLUME]
                                                        # == row[DIVERTED_CONTRAST_VOLUME]
                                                        # == row[LINEAR_DYEVERT_MOVEMENT] == 0
                                                        # and row[DIVERT_VOLUME_DIVERTED] <= 1):
                if row[CUMULATIVE_VOLUME_TO_PATIENT] <= row[THRESHOLD_VOLUME] \
                        / 3 <= row[ATTEMPTED_CONTRAST_INJECTION_VOLUME]:
                    color = LTGRN
                elif row[CUMULATIVE_VOLUME_TO_PATIENT] <= row[THRESHOLD_VOLUME] \
                        * 2 / 3 <= row[ATTEMPTED_CONTRAST_INJECTION_VOLUME]:
                    color = GREEN
                elif row[CUMULATIVE_VOLUME_TO_PATIENT] <= row[THRESHOLD_VOLUME] \
                        <= row[ATTEMPTED_CONTRAST_INJECTION_VOLUME]:
                    color = YELLOW
                elif row[CUMULATIVE_VOLUME_TO_PATIENT] >= row[THRESHOLD_VOLUME] \
                        <= row[ATTEMPTED_CONTRAST_INJECTION_VOLUME]:
                    color = RED
                else:
                    color = WHITE
                cases.append((color, row[DATE_OF_PROCEDURE][0:10], row[DATE_OF_PROCEDURE][11:22],
                              row[THRESHOLD_VOLUME], row[ATTEMPTED_CONTRAST_INJECTION_VOLUME],
                              row[CUMULATIVE_VOLUME_TO_PATIENT], row[DIVERTED_CONTRAST_VOLUME],
                              row[PERCENTAGE_CONTRAST_DIVERTED], uses[row[CMSW_CASE_ID]][1],
                              uses[row[CMSW_CASE_ID]][3], uses[row[CMSW_CASE_ID]][0],
                              uses[row[CMSW_CASE_ID]][2], row[SERIAL_NUMBER]))
    cases.sort(key=_sort_criteria)

    return cases


def dyevert_uses(file_name):
    """Connects to an individual database to calculate the volume of contrast injected
    and the number of times contrast was injected both in puffs and injections
    Volume data is currently unused
    """
    dyevert_used_inj = 0
    dyevert_not_used_inj = 0
    dyevert_used_puff = 0
    dyevert_not_used_puff = 0
    vol_used_inj = 0
    vol_not_used_inj = 0
    vol_used_puff = 0
    vol_not_used_puff = 0
    case_number = 0
    uses = []
    con = sqlite.connect(file_name)
    with con:
        cur = con.cursor()
        cur.execute('SELECT * FROM CMSWCases')
        rows = cur.fetchall()
        number_of_cases = len(rows)
    global TOTAL_DURATION, END_TIME, IS_AN_INJECTION, LINEAR_DYEVERT_MOVEMENT, DYEVERT_CONTRAST_VOLUME_DIVERTED, \
        DYEVERT_CONTRAST_VOLUME_DIVERTED, PERCENT_CONTRAST_SAVED, CONTRAST_VOLUME_TO_PATIENT, \
        PREDOMINANT_CONTRAST_LINE_PRESSURE, FLOW_RATE_TO_FROM_SYRINGE, FLOW_RATE_TO_PATIENT
    if not rows == []:
        if rows[0][2] == '2.1.56' or rows[0][2] == '2.1.24' or rows[0][2] == '2.1.67':
            TOTAL_DURATION = 19
            END_TIME = 20
            IS_AN_INJECTION = 5
            LINEAR_DYEVERT_MOVEMENT = 15
            DYEVERT_CONTRAST_VOLUME_DIVERTED = 17
            PERCENT_CONTRAST_SAVED = 18
            CONTRAST_VOLUME_TO_PATIENT = 20
            FLOW_RATE_TO_FROM_SYRINGE = 28
            FLOW_RATE_TO_PATIENT = 29
            PREDOMINANT_CONTRAST_LINE_PRESSURE = 30
        else:
            TOTAL_DURATION = 20
            END_TIME = 19
            IS_AN_INJECTION = 8
            LINEAR_DYEVERT_MOVEMENT = 18
            DYEVERT_CONTRAST_VOLUME_DIVERTED = 20
            PERCENT_CONTRAST_SAVED = 21
            CONTRAST_VOLUME_TO_PATIENT = 23
            FLOW_RATE_TO_FROM_SYRINGE = 32
            FLOW_RATE_TO_PATIENT = 33
            PREDOMINANT_CONTRAST_LINE_PRESSURE = 34
    with con:

        cur = con.cursor()
        cur.execute('SELECT * FROM CMSWInjections')
        rows = cur.fetchall()

        for row in rows:
            if row[CASE_ID] != case_number:
                uses.append([dyevert_not_used_inj, dyevert_used_inj, dyevert_not_used_puff,
                            dyevert_used_puff])
                dyevert_used_inj = 0
                dyevert_not_used_inj = 0
                dyevert_used_puff = 0
                dyevert_not_used_puff = 0
                vol_used_inj = 0
                vol_not_used_inj = 0
                vol_used_puff = 0
                vol_not_used_puff = 0
                case_number += 1
                while case_number < row[CASE_ID] - 1:
                    uses.append([0, 0, 0, 0, '', ''])
                    case_number += 1
            if row[CASE_ID] == case_number:
                if row[CONTRAST_VOLUME_TO_PATIENT] + row[DYEVERT_CONTRAST_VOLUME_DIVERTED] >= 3:
                    puff_inj = 1
                elif row[CONTRAST_VOLUME_TO_PATIENT] + row[DYEVERT_CONTRAST_VOLUME_DIVERTED] <= 2:
                    puff_inj = 2
                elif row[FLOW_RATE_TO_FROM_SYRINGE] >= 2.5:
                    puff_inj = 1
                elif row[FLOW_RATE_TO_FROM_SYRINGE] <= 2:
                    puff_inj = 2
                if round(row[FLOW_RATE_TO_FROM_SYRINGE], 2) != 0 and round(row[FLOW_RATE_TO_PATIENT], 2) != 0:
                    if row[IS_AN_INJECTION] == 1 and row[PERCENT_CONTRAST_SAVED] == 0 and puff_inj == 1:
                        dyevert_not_used_inj += 1
                        vol_not_used_inj += row[CONTRAST_VOLUME_TO_PATIENT]
                    elif row[IS_AN_INJECTION] == 1 and puff_inj == 1:
                        dyevert_used_inj += 1
                        vol_used_inj += row[CONTRAST_VOLUME_TO_PATIENT]
                    elif row[IS_AN_INJECTION] == 1 and row[PERCENT_CONTRAST_SAVED] == 0 and puff_inj == 2:
                        dyevert_not_used_puff += 1
                        vol_not_used_puff += row[CONTRAST_VOLUME_TO_PATIENT]
                    elif row[IS_AN_INJECTION] == 1 and puff_inj == 2:
                        dyevert_used_puff += 1
                        vol_used_puff += row[CONTRAST_VOLUME_TO_PATIENT]

        uses.append([dyevert_not_used_inj, dyevert_used_inj, dyevert_not_used_puff, dyevert_used_puff])
        while len(uses) <= number_of_cases+1:
            uses.append([0, 0, 0, 0])
        return uses


def excel_write(file_names, cmsw):
    """Writes data into the two Excel Sheets as seen in the example
    Takes two inputs:
        -the file names of the CMSW databases
        -the serial numbers of the CMSWs
    Generates two files:
        -The summary table, which as an augmented sales table
        -The in depth table, which details every injection from the databases
    """
    print('Processing Rod\'s injection data')
    injection_table(file_names, cmsw)
    print('Injection data written, processing summary data')
    cases = list_builder(file_names)
    xlsx1_name = str(cmsw) + 'rods-case-data.xlsx'
    wb = openpyxl.load_workbook('Rods-Template.xlsx')
    data_sheet = wb.active
    data_sheet.title = 'Sheet1'
    print('Writing summary data')
    for row in range(len(cases)):
        for col in range(len(cases[row])):
            data_sheet.cell(row=row + 17, column=col + 1, value=cases[row][col])
            data_sheet.cell(row=row + 17, column=col + 1).alignment = Alignment(wrapText=True)

    data_sheet.column_dimensions['A'].hidden = True
    wb.save(xlsx1_name)
    print('Rod\'s report finished')
